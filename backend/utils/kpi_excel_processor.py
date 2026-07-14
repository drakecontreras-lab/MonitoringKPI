import os
import re
import quopri
import pandas as pd
import numpy as np
import openpyxl
from io import StringIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────

PLANNING_GROUP_MAP = {
    "C81": "Pl.Termoelec.Chuq.",
    "C82": "Pl.Dist.Energ.Elec",
    "C89": "Pl.Mant.Ind.Electr",
    "CO0": "Concentradora A0",
    "CO1": "Concentradora A1",
    "CO2": "Concentradora A2",
    "CB1": "SAE Centralizado",
    "C71": "Pl.Maestranza.Cent",
    "C72": "Pl.Calderería.Cent",
    "C74": "CONTRATOS METALMEC",
    "C40": "Planif. Refinería",
    "SAPCI": "SAPCI"
}

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS GENERALES (compatibles con código existente)
# ──────────────────────────────────────────────────────────────────────────────

def format_proceso_name(name):
    """Normaliza el nombre de proceso quitando prefijos numéricos y sufijos DCH."""
    if pd.isna(name) or name is None:
        return ''
    cleaned = str(name).strip()
    cleaned = re.sub(r'^\d{2}\s+', '', cleaned)
    cleaned = re.sub(r'\s+DCH$', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'DCH$', '', cleaned, flags=re.IGNORECASE)
    if len(cleaned) > 0:
        cleaned = cleaned[0].upper() + cleaned[1:].lower()
    return cleaned


def is_total_or_invalid_row(proceso, gr_planif, gr_planif_pm):
    """Verifica si la fila es un total o tiene datos inválidos."""
    p   = str(proceso or '').strip().lower()
    g   = str(gr_planif or '').strip().lower()
    gpm = str(gr_planif_pm or '').strip().lower()
    invalid_keywords = ['n/a', 'total', 'resultado', 'general', 'código no encontrado',
                        'codigo no encontrado']
    if any(k in p for k in invalid_keywords) or p == '':
        return True
    if any(k in g for k in invalid_keywords) or g == '':
        return True
    if any(k in gpm for k in invalid_keywords) or gpm == '':
        return True
    return False


def get_val(row, idx, default=""):
    """Obtiene el valor de una fila por índice, retorna default si es NaN."""
    try:
        if idx < len(row):
            val = row[idx]
            if pd.isna(val):
                return default
            return val
    except:
        pass
    return default


def get_num(row, idx, default=0.0):
    """Obtiene un valor numérico de una fila por índice."""
    val = get_val(row, idx, default)
    try:
        if str(val).strip() == '':
            return default
        return float(val)
    except:
        return default


def get_safe_proceso(p):
    """Retorna el nombre de proceso normalizado o un valor por defecto."""
    name = str(p or '').strip()
    if not name:
        return 'Sin asignar / Otros'
    return format_proceso_name(name)


def extract_total_percentage(rows, process_col_idx=None):
    """Extrae el porcentaje de cumplimiento de la fila de totales."""
    if not rows or len(rows) <= 1:
        return None
    for i in range(len(rows) - 1, 0, -1):
        row = rows[i]
        is_total_row = False
        for cell in row:
            if pd.notna(cell) and isinstance(cell, str) and ('total' in cell.lower() or 'resultado' in cell.lower()):
                is_total_row = True
                break
        if is_total_row:
            for col_idx in range(len(row) - 1, -1, -1):
                val = get_val(row, col_idx, None)
                if val is not None and isinstance(val, str) and '%' in val:
                    try:
                        cleaned = val.replace('%', '').replace(',', '.').strip()
                        num = float(cleaned)
                        if 0 < num <= 1.0001: return num
                        if 1.0001 < num <= 100: return num / 100
                    except:
                        pass
            for col_idx in range(len(row) - 1, -1, -1):
                val = get_val(row, col_idx, None)
                if val is not None and isinstance(val, (int, float, np.number)):
                    num = float(val)
                    if 0 < num <= 1.0001: return num
            return None
    return None


# ──────────────────────────────────────────────────────────────────────────────
# LECTORES DE ARCHIVOS SAP
# ──────────────────────────────────────────────────────────────────────────────

def read_sap_mime_xls(path):
    """
    Lee un archivo XLS exportado de SAP en formato MIME Multipart HTML.
    Retorna la tabla de datos más grande como DataFrame con columnas numéricas.
    """
    with open(path, 'rb') as f:
        raw = f.read()

    parts = raw.split(b'--NEXTMIME')
    if len(parts) < 2:
        raise ValueError("El archivo no tiene estructura MIME válida.")

    html_raw = parts[1]
    header_end = html_raw.find(b'\r\n\r\n')
    body_qp = html_raw[header_end + 4:]
    decoded = quopri.decodestring(body_qp)
    html_str = decoded.decode('utf-8', errors='replace')

    dfs = pd.read_html(StringIO(html_str))
    if not dfs:
        raise ValueError("No se encontraron tablas HTML en el archivo MIME.")

    # Retornar la tabla con más datos (la tabla principal SAP)
    return max(dfs, key=lambda d: len(d) * len(d.columns))


def read_raw_sap_file(path):
    """
    Detecta el tipo de archivo SAP (MIME-HTML o Excel) y lo lee.
    Retorna DataFrame con headers numéricos (0, 1, 2, ...).
    """
    with open(path, 'rb') as f:
        header_bytes = f.read(10)

    if header_bytes.startswith(b'MIME'):
        return read_sap_mime_xls(path)
    else:
        try:
            return pd.read_excel(path, header=None, engine='openpyxl')
        except Exception:
            return pd.read_excel(path, header=None)


# ──────────────────────────────────────────────────────────────────────────────
# FUNCIONES DE PARSEO DE VALORES SAP
# ──────────────────────────────────────────────────────────────────────────────

def strip_ch01(val):
    """Elimina el prefijo 'CH01/' de un valor SAP."""
    s = str(val or '').strip()
    return s[5:] if s.startswith('CH01/') else s


def clean_pto_trabajo(val):
    """Limpia el puesto de trabajo quitando el prefijo CH01/ y eliminando duplicados si vienen repetidos."""
    s = str(val or '').strip()
    if s.startswith('CH01/'):
        s = s[5:].strip()
    # Si viene repetido como "CGSSERME CGSSERME" o similar
    parts = s.split()
    if len(parts) == 2 and parts[0] == parts[1]:
        s = parts[0]
    return s


def clean_cumplimiento_val(val):
    """Parsea el cumplimiento eliminando decimales y convirtiendo enteros como 10000 a 100."""
    if pd.isna(val) or val is None:
        return None
    s = str(val).strip()
    if s in ('nan', '', '#', 'NaN', 'None', 'Resultado total', 'Resultado'):
        return None
    try:
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        num = float(s)
        if num > 100.01:
            num = num / 100.0
        return int(round(num))
    except Exception:
        return val



def parse_sap_hh(val):
    """
    Parsea valores de HH exportados por SAP.
    SAP exporta sin separador decimal: '22000' = 22.000 HH → se divide entre 1000.
    Formato europeo: '1.304,600' = 1304.6 HH.
    """
    s = str(val).strip()
    if s in ('nan', '', '#', 'NaN', 'None', 'Resultado total', 'Resultado'):
        return 0.0
    try:
        if ',' in s:
            # Formato europeo: '1.304,600' → 1304.6
            s = s.replace('.', '').replace(',', '.')
            return float(s)
        else:
            v = float(s)
            # SAP exporta HH multiplicadas por 1000
            return v / 1000.0
    except Exception:
        return 0.0


def parse_sap_count(val):
    """Parsea un conteo entero exportado por SAP (valor literal)."""
    s = str(val).strip()
    if s in ('nan', '', '#', 'NaN', 'None', 'Resultado total', 'Resultado'):
        return 0
    try:
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        return int(float(s))
    except Exception:
        return 0


def parse_sap_count_div1000(val):
    """
    Parsea un conteo exportado por SAP multiplicado x1000.
    Ej: '1000' → 1 operación, '391000' → 391 operaciones.
    """
    return parse_sap_count(val) // 1000


def is_resultado_row(row):
    """Verifica si la fila es una fila de total/subtotal SAP ('Resultado' o 'Resultado total')."""
    for val in list(row)[:10]:
        s = str(val).strip()
        if s in ('Resultado', 'Resultado total'):
            return True
    return False


def is_resultado_intermedio(row):
    """
    Verifica si la fila es un subtotal intermedio de SAP ('Resultado').
    Las filas intermedias tienen exactamente 'Resultado' (no 'Resultado total').
    """
    for val in list(row)[:10]:
        s = str(val).strip()
        if s == 'Resultado':
            return True
    return False


def filtrar_filas_resultado_intermedias(df):
    """
    Elimina filas de subtotal intermedio SAP ('Resultado'), conservando solo laúltima fila 'Resultado total'.
    """
    mask = ~df.apply(lambda r: is_resultado_intermedio(r), axis=1)
    return df[mask]


def aplicar_formato_tabla_openpyxl(ws, sheet_name):
    """
    Aplica formato tabla de Excel a la hoja dada en un Workbook de openpyxl.
    El nombre de tabla se deriva del nombre de la hoja.
    """
    try:
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 2 or max_col < 1:
            return
        # Crear nombre de tabla limpio (sin espacios ni caracteres especiales)
        nombre_tabla = re.sub(r'[^A-Za-z0-9]', '_', sheet_name)
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tabla = Table(displayName=nombre_tabla, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tabla.tableStyleInfo = style
        ws.add_table(tabla)
    except Exception as e:
        print(f"[ExcelProcessor] Advertencia al aplicar formato tabla en {sheet_name}: {e}")


def _make_unique_headers(headers):
    """Asegura que los nombres de columna sean únicos añadiendo sufijos cuando hay duplicados."""
    seen = {}
    result = []
    for h in headers:
        h_str = str(h).strip() if str(h).strip() not in ('nan', '') else 'Col'
        if h_str in seen:
            seen[h_str] += 1
            result.append(f"{h_str}_{seen[h_str]}")
        else:
            seen[h_str] = 0
            result.append(h_str)
    return result


# ──────────────────────────────────────────────────────────────────────────────
# EXTRACTORES POR TIPO DE ARCHIVO
# ──────────────────────────────────────────────────────────────────────────────

def extract_avisos(path, puestos_mapping=None, grupos_mapping=None):
    """
    Procesa el archivo de Avisos Pendientes (SAP MIME-HTML o XLSX).
    Estructura del archivo:
      - fila 0 (iloc[0]): headers SAP
      - fila 1+ (iloc[1:-1]): datos (col[2]=Proceso, col[10]=Gr.planif.PM, col[11]=CH01/código)
      - última fila: 'Resultado total'
    Retorna (df_limpio, stats_dict).
    """
    if grupos_mapping is None:
        grupos_mapping = dict(PLANNING_GROUP_MAP)
    df = read_raw_sap_file(path)

    group = {}
    total = 0

    data_rows = df.iloc[1:-1]

    for _, row in data_rows.iterrows():
        row_list = list(row)
        if is_resultado_row(row_list):
            continue

        proceso_raw   = str(row_list[2]  if len(row_list) > 2  else '').strip()
        gr_planif_pm  = str(row_list[10] if len(row_list) > 10 else '').strip()
        gr_planif_raw = str(row_list[11] if len(row_list) > 11 else '').strip()
        pto_trabajo_raw = str(row_list[15] if len(row_list) > 15 else '').strip()
        pto_trabajo = clean_pto_trabajo(pto_trabajo_raw) if pto_trabajo_raw else 'N/A'
        gr_planif     = strip_ch01(gr_planif_raw)

        if not proceso_raw or proceso_raw == 'nan':
            continue

        proceso = get_safe_proceso(proceso_raw)
        if not gr_planif or gr_planif in ('nan', '#'):
            gr_planif = 'N/A'
        if not gr_planif_pm or gr_planif_pm == 'nan':
            gr_planif_pm = grupos_mapping.get(gr_planif, gr_planif or 'N/A')

        key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
        group[key] = group.get(key, 0) + 1
        total += 1

    # Construir DataFrame limpio para exportación al XLSX consolidado
    df_clean = data_rows.copy()

    # Aplicar headers del archivo SAP (fila 0) y renombrar columnas clave
    headers = _make_unique_headers(list(df.iloc[0]))
    rename_map = {5: 'Txt. breve', 11: 'Gr. Planif', 13: 'Denom', 15: 'Pto. Trabajo'}
    for idx, name in rename_map.items():
        if idx < len(headers):
            headers[idx] = name
    df_clean.columns = headers

    # Limpiar prefijo CH01/ en columnas Gr. Planif y aplicar clean_pto_trabajo a Pto. Trabajo
    for col_name in ('Gr. Planif', 'Gr. planif', 'Gr.planif'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].astype(str).str.replace('CH01/', '', regex=False)
    for col_name in ('Pto. Trabajo', 'Pto. Trabajo Descripcion'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].apply(clean_pto_trabajo)
    if puestos_mapping and 'Pto. Trabajo Descripcion' in df_clean.columns and 'Pto. Trabajo' in df_clean.columns:
        df_clean['Pto. Trabajo Descripcion'] = df_clean['Pto. Trabajo'].apply(lambda x: str(puestos_mapping.get(x)).capitalize() if puestos_mapping.get(x) else 'N/A')

    if 'Pto. Trabajo' in df_clean.columns and 'Pto. Trabajo Descripcion' in df_clean.columns:
        cols = list(df_clean.columns)
        cols.remove('Pto. Trabajo')
        idx = cols.index('Pto. Trabajo Descripcion')
        cols.insert(idx + 1, 'Pto. Trabajo')
        df_clean = df_clean[cols]

    df_clean = df_clean.where(pd.notnull(df_clean), None)

    distribucion = []
    for key in sorted(group.keys()):
        parts = key.split('||')
        p, gp, gppm = parts[0], parts[1], parts[2]
        pto = parts[3] if len(parts) > 3 else 'N/A'
        pto_desc = str(puestos_mapping.get(pto, 'N/A')).capitalize() if puestos_mapping and puestos_mapping.get(pto) else 'N/A'
        distribucion.append({'proceso': p, 'grPlanif': gp, 'grPlanifPM': gppm, 'ptoTrabajo': pto, 'ptoTrabajoDesc': pto_desc, 'cantidad': group[key]})

    return df_clean, {'total': total, 'distribucion': distribucion}


def extract_ordenes(path, puestos_mapping=None, grupos_mapping=None):
    """
    Procesa el archivo de Órdenes Pendientes (SAP MIME-HTML o XLSX).
    Estructura:
      - fila 0: headers SAP (col[0]=Gr.planif.PM, col[1]=CH01/código, col[3]=Proceso)
      - datos en fila 1:-1
    Retorna (df_limpio, stats_dict).
    """
    if grupos_mapping is None:
        grupos_mapping = dict(PLANNING_GROUP_MAP)
    df = read_raw_sap_file(path)

    group = {}
    total = 0

    data_rows = df.iloc[1:-1]

    gr_planif_col = []
    gr_planif_pm_col = []

    for _, row in data_rows.iterrows():
        row_list = list(row)
        if is_resultado_row(row_list):
            gr_planif_col.append(None)
            gr_planif_pm_col.append(None)
            continue

        gr_planif_pm  = str(row_list[0] if len(row_list) > 0 else '').strip()
        gr_planif_raw = str(row_list[1] if len(row_list) > 1 else '').strip()
        gr_planif     = strip_ch01(gr_planif_raw)
        proceso_raw   = str(row_list[3] if len(row_list) > 3 else '').strip()
        pto_trabajo_raw = str(row_list[13] if len(row_list) > 13 else '').strip()
        pto_trabajo = clean_pto_trabajo(pto_trabajo_raw) if pto_trabajo_raw else 'N/A'

        if not proceso_raw or proceso_raw == 'nan':
            gr_planif_col.append('N/A')
            gr_planif_pm_col.append('N/A')
            continue

        proceso = get_safe_proceso(proceso_raw)
        if not gr_planif or gr_planif in ('nan', '#'):
            gr_planif = 'N/A'
        if not gr_planif_pm or gr_planif_pm == 'nan':
            gr_planif_pm = grupos_mapping.get(gr_planif, gr_planif or 'N/A')

        gr_planif_col.append(gr_planif)
        gr_planif_pm_col.append(gr_planif_pm)

        key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
        group[key] = group.get(key, 0) + 1
        total += 1

    # DataFrame limpio
    df_clean = data_rows.copy()
    headers = _make_unique_headers(list(df.iloc[0]))
    rename_map = {1: 'Gr. Planif', 6: 'Txt. breve', 12: 'Pto. Trabajo Descripcion', 13: 'Pto. Trabajo'}
    for idx, name in rename_map.items():
        if idx < len(headers):
            headers[idx] = name
    df_clean.columns = headers

    # Sobrescribir posicionalmente col0 (Gr.planif.PM nativo) y col1 (Gr. Planif renombrada)
    # con el valor enriquecido (nunca vacío). Sin crear columnas nuevas: esos headers
    # ya existen en el archivo crudo de SAP y duplicarlos corrompe la tabla Excel.
    if len(df_clean) == len(gr_planif_col):
        if len(df_clean.columns) > 1:
            df_clean.iloc[:, 1] = gr_planif_col
        if len(df_clean.columns) > 0:
            df_clean.iloc[:, 0] = gr_planif_pm_col

    for col_name in ('Gr. Planif', 'Gr. planif', 'Gr.planif'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].astype(str).str.replace('CH01/', '', regex=False)
    for col_name in ('Pto. Trabajo', 'Pto. Trabajo Descripcion'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].apply(clean_pto_trabajo)
    if puestos_mapping and 'Pto. Trabajo Descripcion' in df_clean.columns and 'Pto. Trabajo' in df_clean.columns:
        df_clean['Pto. Trabajo Descripcion'] = df_clean['Pto. Trabajo'].apply(lambda x: str(puestos_mapping.get(x)).capitalize() if puestos_mapping.get(x) else 'N/A')

    if 'Pto. Trabajo' in df_clean.columns and 'Pto. Trabajo Descripcion' in df_clean.columns:
        cols = list(df_clean.columns)
        cols.remove('Pto. Trabajo')
        idx = cols.index('Pto. Trabajo Descripcion')
        cols.insert(idx + 1, 'Pto. Trabajo')
        df_clean = df_clean[cols]

    df_clean = df_clean.where(pd.notnull(df_clean), None)

    distribucion = []
    for key in sorted(group.keys()):
        parts = key.split('||')
        p, gp, gppm = parts[0], parts[1], parts[2]
        pto = parts[3] if len(parts) > 3 else 'N/A'
        pto_desc = str(puestos_mapping.get(pto, 'N/A')).capitalize() if puestos_mapping and puestos_mapping.get(pto) else 'N/A'
        distribucion.append({'proceso': p, 'grPlanif': gp, 'grPlanifPM': gppm, 'ptoTrabajo': pto, 'ptoTrabajoDesc': pto_desc, 'cantidad': group[key]})

    return df_clean, {'total': total, 'distribucion': distribucion}


def extract_trabajo_planificado(path, ots_mapping=None, puestos_mapping=None, grupos_mapping=None):
    """
    Procesa el archivo de % Trabajo Planificado (SAP MIME-HTML o XLSX).
    Propósito: Estructurar y limpiar los datos de Trabajo Planificado.
    - Primero intenta leer gr_planif/gr_planif_pm directamente de las columnas del archivo PATTERN (cols 4/5).
    - Si esas columnas están vacías, las complementa con ots_mapping (proveniente del EXPORT IW39).
    """
    if ots_mapping is None:
        ots_mapping = {}
    if grupos_mapping is None:
        grupos_mapping = dict(PLANNING_GROUP_MAP)

    df = read_raw_sap_file(path)

    group = {}
    total_planificado = 0.0
    total_sin_hr      = 0.0
    total_imprevistos = 0.0

    data_rows = df.iloc[4:-1]  # Saltar 4 filas de encabezado (title, empty, partial, full headers) + última de total

    criterios_col = []
    gr_planif_col = []
    gr_planif_pm_col = []

    for _, row in data_rows.iterrows():
        row_list = list(row)
        if is_resultado_row(row_list):
            criterios_col.append('')
            gr_planif_col.append(None)
            gr_planif_pm_col.append(None)
            continue

        proceso_raw  = str(row_list[2]  if len(row_list) > 2  else '').strip()
        orden        = str(row_list[8]  if len(row_list) > 8  else '').strip()
        clase_orden  = str(row_list[11] if len(row_list) > 11 else '').strip()
        fecha_libera = str(row_list[12] if len(row_list) > 12 else '').strip()
        fecha_inicio = str(row_list[13] if len(row_list) > 13 else '').strip()
        pto_trabajo_raw = str(row_list[5] if len(row_list) > 5 else '').strip()
        pto_trabajo = clean_pto_trabajo(pto_trabajo_raw) if pto_trabajo_raw else 'N/A'
        grupo_ruta   = str(row_list[14] if len(row_list) > 14 else '').strip()
        pct_planif   = parse_sap_count(row_list[15] if len(row_list) > 15 else 0)  # 10000 = 100%
        hh_totales   = parse_sap_hh(row_list[17] if len(row_list) > 17 else 0)

        if not proceso_raw or proceso_raw == 'nan':
            criterios_col.append('')
            gr_planif_col.append('N/A')
            gr_planif_pm_col.append('N/A')
            continue

        proceso = get_safe_proceso(proceso_raw)

        # Paso 1: intentar leer gr_planif/gr_planif_pm directamente del PATTERN (cols 4 y 5)
        # El PATTERN puede tener distintos layouts; col 4 = nombre PM, col 5 = código con/sin CH01/
        gr_planif_pm_raw = str(row_list[4] if len(row_list) > 4 else '').strip()
        gr_planif_raw_col = str(row_list[5] if len(row_list) > 5 else '').strip()
        gr_planif_from_file = strip_ch01(gr_planif_raw_col)

        # Verificar si col 4 parece un nombre de grupo PM (no es puesto de trabajo)
        # Heurística: los nombres de grupo PM del mapa son valores cortos sin espacios o tipo 'Pl.XXX'
        name_is_pm = any(
            v.lower() in gr_planif_pm_raw.lower()
            for v in grupos_mapping.values()
        )
        # Verificar si col 5 parece un código de grupo (C81, C82, CO0, etc.)
        code_is_planif = gr_planif_from_file in grupos_mapping

        if code_is_planif:
            # El archivo tiene el código directamente → usar columnas del PATTERN
            gr_planif    = gr_planif_from_file
            gr_planif_pm = grupos_mapping.get(gr_planif, gr_planif_pm_raw if name_is_pm else gr_planif)
        elif name_is_pm:
            # Sólo el nombre PM está disponible → deducir código desde el mapa inverso
            inv_map = {v: k for k, v in grupos_mapping.items()}
            gr_planif    = inv_map.get(gr_planif_pm_raw, 'N/A')
            gr_planif_pm = gr_planif_pm_raw
        else:
            # Paso 2: columnas del PATTERN no tienen datos de grupo → usar ots_mapping
            gr_planif    = 'N/A'
            gr_planif_pm = 'N/A'

        # Complementar con ots_mapping si todavía está vacío
        # Normalizar orden: SAP puede exportarla como '110415552.0' (float-string)
        # mientras ots_mapping usa claves int-string '110415552' (ver main.py,
        # construcción de ots_mapping); sin esto el lookup fallaba siempre.
        try:
            orden_norm = str(int(float(orden))) if orden and orden not in ('nan', '') else orden
        except ValueError:
            orden_norm = orden

        if (gr_planif in ('N/A', '', 'nan') or gr_planif_pm in ('N/A', '', 'nan')) and orden_norm in ots_mapping:
            mapped_gp, mapped_gppm = ots_mapping[orden_norm]
            if gr_planif in ('N/A', '', 'nan'):
                gr_planif = mapped_gp
            if gr_planif_pm in ('N/A', '', 'nan'):
                gr_planif_pm = mapped_gppm

        gr_planif_col.append(gr_planif)
        gr_planif_pm_col.append(gr_planif_pm)

        # ── Lógica de criterios ───────────────────────────────────────────────
        # Reglas:
        #   1. NP (o clase no PL)       → Imprevisto
        #   2. PL + grupo_ruta = '#'/'' → Sin HR
        #   3. PL + grupo_ruta válido
        #        + % planificado = 100% (pct_planif >= 9900, SAP exporta 10000 = 100%)
        #        + (Fecha inicio real - Fecha Libera Orden) > 7 días
        #                                → Planificado
        #   4. PL que no cumple (3)     → Sin horizonte

        if clase_orden != 'PL':
            criterio = 'Imprevisto'
        elif grupo_ruta in ('#', '', 'nan'):
            criterio = 'Sin HR'
        else:
            # Calcular diferencia de fechas (formato SAP: 'DD.MM.YYYY')
            dias_horizonte = None
            try:
                from datetime import datetime
                fmt = '%d.%m.%Y'
                f_inicio  = datetime.strptime(fecha_inicio,  fmt)
                f_libera  = datetime.strptime(fecha_libera,  fmt)
                dias_horizonte = (f_inicio - f_libera).days
            except Exception:
                dias_horizonte = None  # fecha inválida o vacía

            cumple_pct     = pct_planif >= 9900                          # 100% cumplimiento
            cumple_horizon = (dias_horizonte is not None and dias_horizonte > 7)

            if cumple_pct and cumple_horizon:
                criterio = 'Planificado'
            else:
                criterio = 'Sin horizonte'

        criterios_col.append(criterio)

        key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
        if key not in group:
            group[key] = {'planificado': 0.0, 'sinHr': 0.0, 'sinHorizonte': 0.0, 'imprevistos': 0.0}

        if criterio == 'Planificado':
            group[key]['planificado'] += hh_totales
            total_planificado += hh_totales
        elif criterio == 'Sin HR':
            group[key]['sinHr'] += hh_totales
            total_sin_hr += hh_totales
        elif criterio == 'Sin horizonte':
            group[key]['sinHorizonte'] += hh_totales
            total_sin_hr += hh_totales  # ambos descuentan del cumplimiento
        else:
            group[key]['imprevistos'] += hh_totales
            total_imprevistos += hh_totales

    total_total   = total_planificado + total_sin_hr + total_imprevistos
    cumplimiento  = total_planificado / total_total if total_total > 0 else 0.0

    grupos = []
    for key in sorted(group.keys()):
        parts = key.split('||')
        p, gp, gppm = parts[0], parts[1], parts[2]
        pto = parts[3] if len(parts) > 3 else 'N/A'
        pto_desc = str(puestos_mapping.get(pto, 'N/A')).capitalize() if puestos_mapping and puestos_mapping.get(pto) else 'N/A'
        vals = group[key]
        rt   = vals['planificado'] + vals['sinHr'] + vals.get('sinHorizonte', 0.0) + vals['imprevistos']
        c    = vals['planificado'] / rt if rt > 0 else 0.0
        grupos.append({
            'proceso': p, 'grPlanif': gp, 'grPlanifPM': gppm, 'ptoTrabajo': pto, 'ptoTrabajoDesc': pto_desc,
            'planificado': vals['planificado'],
            'sinHr': vals['sinHr'],
            'sinHorizonte': vals.get('sinHorizonte', 0.0),
            'imprevistos': vals['imprevistos'],
            'total': rt, 'cumplimiento': c
        })

    # Construir DataFrame limpio para exportación
    df_clean = data_rows.copy()

    # Encabezados: combinar fila 0 (cols numéricas) y fila 1 (texto)
    h0 = list(df.iloc[0])
    h1 = list(df.iloc[1])
    merged = []
    for i, (a, b) in enumerate(zip(h0, h1)):
        a_s, b_s = str(a).strip(), str(b).strip()
        if a_s not in ('nan', ''):
            merged.append(a_s)
        elif b_s not in ('nan', ''):
            merged.append(b_s)
        else:
            merged.append(f'Col_{i}')
    merged = _make_unique_headers(merged)
    df_clean.columns = merged

    # Renombrar explícitamente col 15 a % Trabajo Planificado y 4 y 5 a Puesto de Trabajo
    if len(df_clean.columns) > 15:
        cols = list(df_clean.columns)
        cols[15] = '% Trabajo Planificado'
        cols[4] = 'Pto. Trabajo Descripcion'
        cols[5] = 'Pto. Trabajo'
        df_clean.columns = cols

    # Asignar Criterio y añadir los grupos calculados como nuevas columnas
    if len(df_clean) == len(criterios_col):
        df_clean['Criterio'] = criterios_col
        df_clean['Gr.planif.PM'] = gr_planif_pm_col
        df_clean['Gr.planif'] = gr_planif_col

    # Limpiar columnas
    for col_name in ('Gr.planif.PM', 'Gr.planif', 'Gr. Planif', 'Gr. planif'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].astype(str).str.replace('CH01/', '', regex=False)
    for col_name in ('Pto. Trabajo', 'Pto. Trabajo Descripcion'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].apply(clean_pto_trabajo)
    if '% Trabajo Planificado' in df_clean.columns:
        df_clean['% Trabajo Planificado'] = df_clean['% Trabajo Planificado'].apply(clean_cumplimiento_val)
    if puestos_mapping and 'Pto. Trabajo Descripcion' in df_clean.columns and 'Pto. Trabajo' in df_clean.columns:
        df_clean['Pto. Trabajo Descripcion'] = df_clean['Pto. Trabajo'].apply(lambda x: str(puestos_mapping.get(x)).capitalize() if puestos_mapping.get(x) else 'N/A')

    if 'Pto. Trabajo' in df_clean.columns and 'Pto. Trabajo Descripcion' in df_clean.columns:
        cols = list(df_clean.columns)
        cols.remove('Pto. Trabajo')
        idx = cols.index('Pto. Trabajo Descripcion')
        cols.insert(idx + 1, 'Pto. Trabajo')
        df_clean = df_clean[cols]

    # FIX: HH Plan. Reales y HH Totales Reales vienen de SAP x1000 (22000 = 22.000 HH)
    for col_hh in ('HH Plan. Reales', 'HH Totales Reales'):
        if col_hh in df_clean.columns:
            df_clean[col_hh] = pd.to_numeric(df_clean[col_hh], errors='coerce').fillna(0) / 1000.0

    df_clean = df_clean.where(pd.notnull(df_clean), None)

    return df_clean, {
        'grupos': grupos,
        'total': {
            'planificado': total_planificado,
            'sinHr': total_sin_hr,
            'imprevistos': total_imprevistos,
            'total': total_total,
            'cumplimiento': cumplimiento
        }
    }



def extract_programa_semanal(path, puestos_mapping=None, grupos_mapping=None):
    """
    Procesa el archivo de Programa Semanal (SAP MIME-HTML o XLSX).
    Estructura (doble encabezado):
      - fila 0: parcial (cols 14-16: Cumplimiento%, Ind.cumple, Total Op)
      - fila 1: texto completo (col[0]=Gr.planif.PM, col[1]=CH01/código, col[4]=Proceso)
      - fila 2+: datos
    Criterio: Cumplimiento (col[14]) = 10000 (100.00%) → Cumple, sino → No cumple.
    KPI = Σ(Ind.cumple) / Σ(Total Op ÷ 1000).
    Retorna (df_limpio, stats_dict).
    """
    if grupos_mapping is None:
        grupos_mapping = dict(PLANNING_GROUP_MAP)
    df = read_raw_sap_file(path)

    group = {}
    total_ind_cumple = 0.0
    total_ops        = 0.0

    data_rows   = df.iloc[2:-1]
    criterios_col = []
    gr_planif_col = []
    gr_planif_pm_col = []

    for _, row in data_rows.iterrows():
        row_list = list(row)
        if is_resultado_row(row_list):
            criterios_col.append('')
            gr_planif_col.append('')
            gr_planif_pm_col.append('')
            continue

        gr_planif_pm  = str(row_list[0] if len(row_list) > 0 else '').strip()
        gr_planif_raw = str(row_list[1] if len(row_list) > 1 else '').strip()
        gr_planif     = strip_ch01(gr_planif_raw)
        proceso_raw   = str(row_list[4] if len(row_list) > 4 else '').strip()
        pto_trabajo_raw = str(row_list[7] if len(row_list) > 7 else '').strip()
        pto_trabajo = clean_pto_trabajo(pto_trabajo_raw) if pto_trabajo_raw else 'N/A'

        if not proceso_raw or proceso_raw == 'nan':
            criterios_col.append('')
            gr_planif_col.append('N/A')
            gr_planif_pm_col.append('N/A')
            continue

        cumpl_val  = parse_sap_count(row_list[14] if len(row_list) > 14 else 0)
        ind_cumple = parse_sap_count(row_list[15] if len(row_list) > 15 else 0)
        total_op   = parse_sap_count_div1000(row_list[16] if len(row_list) > 16 else 0)

        proceso = get_safe_proceso(proceso_raw)
        if not gr_planif or gr_planif in ('nan', '#'):
            gr_planif = 'N/A'
        if not gr_planif_pm or gr_planif_pm == 'nan':
            gr_planif_pm = grupos_mapping.get(gr_planif, gr_planif or 'N/A')

        # Cumple si el porcentaje es 100% (SAP lo exporta como 10000 = 100.00%)
        cumple_flag = cumpl_val >= 9900
        criterio = 'Cumple' if cumple_flag else 'No cumple'
        criterios_col.append(criterio)
        gr_planif_col.append(gr_planif)
        gr_planif_pm_col.append(gr_planif_pm)

        key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
        if key not in group:
            group[key] = {'cumple': 0.0, 'noCumple': 0.0, 'sumIndCumple': 0.0, 'sumTotalOp': 0.0}

        group[key]['sumIndCumple'] += ind_cumple
        group[key]['sumTotalOp']   += total_op
        if cumple_flag:
            group[key]['cumple'] += total_op
        else:
            group[key]['noCumple'] += total_op

        total_ind_cumple += ind_cumple
        total_ops        += total_op

    cumplimiento_total = total_ind_cumple / total_ops if total_ops > 0 else 0.0

    grupos = []
    for key in sorted(group.keys()):
        parts = key.split('||')
        p, gp, gppm = parts[0], parts[1], parts[2]
        pto = parts[3] if len(parts) > 3 else 'N/A'
        pto_desc = str(puestos_mapping.get(pto, 'N/A')).capitalize() if puestos_mapping and puestos_mapping.get(pto) else 'N/A'
        vals = group[key]
        cump = vals['sumIndCumple'] / vals['sumTotalOp'] if vals['sumTotalOp'] > 0 else 0.0
        grupos.append({
            'proceso': p, 'grPlanif': gp, 'grPlanifPM': gppm, 'ptoTrabajo': pto, 'ptoTrabajoDesc': pto_desc,
            'cumple': vals['cumple'], 'noCumple': vals['noCumple'],
            'total': vals['sumTotalOp'], 'cumplimiento': cump
        })

    # DataFrame limpio
    df_clean = data_rows.copy()
    h0 = list(df.iloc[0])
    h1 = list(df.iloc[1])
    merged = []
    for i, (a, b) in enumerate(zip(h0, h1)):
        a_s, b_s = str(a).strip(), str(b).strip()
        if a_s not in ('nan', ''):
            merged.append(a_s)
        elif b_s not in ('nan', ''):
            merged.append(b_s)
        else:
            merged.append(f'Col_{i}')
    merged = _make_unique_headers(merged)
    df_clean.columns = merged

    # Renombrar explícitamente columnas
    if len(df_clean.columns) > 7:
        cols = list(df_clean.columns)
        cols[5] = 'Pto. Trabajo Descripcion'
        cols[7] = 'Pto. Trabajo'
        df_clean.columns = cols

    # Limpiar columnas de grupo y puesto de trabajo
    for col_name in ('Gr. Planif', 'Gr. planif', 'Gr.planif'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].astype(str).str.replace('CH01/', '', regex=False)
    for col_name in ('Pto. Trabajo', 'Pto. Trabajo Descripcion'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].apply(clean_pto_trabajo)

    # Limpiar cumplimiento (columna 14)
    if len(df_clean.columns) > 14:
        col_cumpl = df_clean.columns[14]
        df_clean[col_cumpl] = df_clean[col_cumpl].apply(clean_cumplimiento_val)

    df_clean = df_clean.copy()
    df_clean['Criterio'] = criterios_col
    df_clean['Gr.planif'] = gr_planif_col
    df_clean['Gr.planif.PM'] = gr_planif_pm_col
    if puestos_mapping and 'Pto. Trabajo Descripcion' in df_clean.columns and 'Pto. Trabajo' in df_clean.columns:
        df_clean['Pto. Trabajo Descripcion'] = df_clean['Pto. Trabajo'].apply(lambda x: str(puestos_mapping.get(x)).capitalize() if puestos_mapping.get(x) else 'N/A')

    if 'Pto. Trabajo' in df_clean.columns and 'Pto. Trabajo Descripcion' in df_clean.columns:
        cols = list(df_clean.columns)
        cols.remove('Pto. Trabajo')
        idx = cols.index('Pto. Trabajo Descripcion')
        cols.insert(idx + 1, 'Pto. Trabajo')
        df_clean = df_clean[cols]

    # FIX: Total Op. Programadas viene x1000 desde SAP (1000 = 1 operacion)
    if 'Total Op. Programadas' in df_clean.columns:
        df_clean['Total Op. Programadas'] = pd.to_numeric(df_clean['Total Op. Programadas'], errors='coerce').fillna(0) / 1000.0

    df_clean = df_clean.where(pd.notnull(df_clean), None)

    return df_clean, {
        'grupos': grupos,
        'total': {
            'cumple': sum(g['cumple'] for g in grupos),
            'noCumple': sum(g['noCumple'] for g in grupos),
            'total': total_ops,
            'cumplimiento': cumplimiento_total
        }
    }




def extract_plan_matriz(path, export_ops_mapping=None, puestos_mapping=None, grupos_mapping=None):
    """
    Procesa el archivo de Plan Matriz (SAP MIME-HTML o XLSX).
    Estructura (doble encabezado):
      - fila 0: parcial (cols 14-18: %Cumpl, PlanMatriz, CoefPMA, Op.Ejec, Op.Total)
      - fila 1: texto completo (col[2]=Proceso, col[4]=Gr.planif.PM, col[5]=CH01/código)
      - fila 2+: datos
    Criterio: Op.Ejec (col[17]) >= Op.Total (col[18]) → Cumple.
    KPI = Σ(Op.Ejec) / Σ(Op.Total).
    - Los valores de op_ejec y op_total vienen en escala real (no multiplicados ×1000).
    - Si se provee export_ops_mapping {orden: cant_ops_reales}, corrige op_total con ese conteo.
    Retorna (df_limpio, stats_dict).
    """
    if export_ops_mapping is None:
        export_ops_mapping = {}
    if grupos_mapping is None:
        grupos_mapping = dict(PLANNING_GROUP_MAP)

    df = read_raw_sap_file(path)

    group = {}
    total_ejec    = 0.0
    total_totales = 0.0

    data_rows     = df.iloc[2:-1]
    criterios_col = []
    gr_planif_col = []
    gr_planif_pm_col = []
    nuevos_totales_col = []
    nuevos_ejec_col = []

    # Detectar columna de orden en el encabezado (fila 1)
    col_orden = 8
    if len(df) > 1:
        header_row = [str(v).strip() for v in list(df.iloc[1])]
        for i, h in enumerate(header_row):
            if 'orden' in h.lower() and 'mantenimiento' in h.lower():
                col_orden = i
                break

    for _, row in data_rows.iterrows():
        row_list = list(row)
        if is_resultado_row(row_list):
            criterios_col.append('')
            gr_planif_col.append('')
            gr_planif_pm_col.append('')
            nuevos_totales_col.append(None)
            nuevos_ejec_col.append(None)
            continue

        gr_planif_pm  = str(row_list[4] if len(row_list) > 4 else '').strip()
        gr_planif_raw = str(row_list[5] if len(row_list) > 5 else '').strip()
        gr_planif     = strip_ch01(gr_planif_raw)
        proceso_raw   = str(row_list[2] if len(row_list) > 2 else '').strip()
        pto_trabajo_raw = str(row_list[13] if len(row_list) > 13 else '').strip()
        pto_trabajo = clean_pto_trabajo(pto_trabajo_raw) if pto_trabajo_raw else 'N/A'

        if not proceso_raw or proceso_raw == 'nan':
            criterios_col.append('')
            gr_planif_col.append('N/A')
            gr_planif_pm_col.append('N/A')
            nuevos_totales_col.append(None)
            nuevos_ejec_col.append(None)
            continue

        # Los valores del PATTERN están en escala real (no × 1000)
        op_ejec = parse_sap_count(row_list[17] if len(row_list) > 17 else 0)
        # op_total base desde la columna 18 del PATTERN (escala real)
        op_total_base = parse_sap_count(row_list[18] if len(row_list) > 18 else 0)
        nuevos_ejec_col.append(op_ejec)

        # Obtener orden para buscar en export_ops_mapping
        orden = str(row_list[col_orden] if len(row_list) > col_orden else '').strip()
        # Normalizar: quitar decimales si el número viene como float string (ej. '110415552.0')
        try:
            orden_norm = str(int(float(orden))) if orden and orden not in ('nan', '') else orden
        except ValueError:
            orden_norm = orden

        # Corrección de op_total: usar conteo real del EXPORT IW37N si está disponible
        if orden_norm in export_ops_mapping:
            op_total = export_ops_mapping[orden_norm]
        else:
            op_total = op_total_base

        nuevos_totales_col.append(op_total)

        proceso = get_safe_proceso(proceso_raw)
        if not gr_planif or gr_planif in ('nan', '#'):
            gr_planif = 'N/A'
        if not gr_planif_pm or gr_planif_pm == 'nan':
            gr_planif_pm = grupos_mapping.get(gr_planif, gr_planif or 'N/A')





        # Cumple si ejecutó todo lo planificado
        cumple_flag = (op_ejec >= op_total and op_total > 0)
        criterio = 'Cumple' if cumple_flag else 'No cumple'
        criterios_col.append(criterio)
        gr_planif_col.append(gr_planif)
        gr_planif_pm_col.append(gr_planif_pm)

        key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
        if key not in group:
            group[key] = {'ejec': 0.0, 'total_plan': 0.0}

        group[key]['ejec']       += op_ejec
        group[key]['total_plan'] += op_total
        total_ejec    += op_ejec
        total_totales += op_total

    cumplimiento_total = total_ejec / total_totales if total_totales > 0 else 0.0

    grupos = []
    for key in sorted(group.keys()):
        parts = key.split('||')
        p, gp, gppm = parts[0], parts[1], parts[2]
        pto = parts[3] if len(parts) > 3 else 'N/A'
        pto_desc = str(puestos_mapping.get(pto, 'N/A')).capitalize() if puestos_mapping and puestos_mapping.get(pto) else 'N/A'
        vals = group[key]
        cump = vals['ejec'] / vals['total_plan'] if vals['total_plan'] > 0 else 0.0
        grupos.append({
            'proceso': p, 'grPlanif': gp, 'grPlanifPM': gppm, 'ptoTrabajo': pto, 'ptoTrabajoDesc': pto_desc,
            'cumple': vals['ejec'],
            'noCumple': vals['total_plan'] - vals['ejec'],
            'total': vals['total_plan'],
            'cumplimiento': cump
        })

    # DataFrame limpio
    df_clean = data_rows.copy()
    h0 = list(df.iloc[0])
    h1 = list(df.iloc[1])
    merged = []
    for i, (a, b) in enumerate(zip(h0, h1)):
        a_s, b_s = str(a).strip(), str(b).strip()
        if a_s not in ('nan', ''):
            merged.append(a_s)
        elif b_s not in ('nan', ''):
            merged.append(b_s)
        else:
            merged.append(f'Col_{i}')
    merged = _make_unique_headers(merged)
    df_clean.columns = merged

    # Renombrar columnas 12 y 13 para Puesto de Trabajo
    if len(df_clean.columns) > 13:
        cols = list(df_clean.columns)
        cols[12] = 'Pto. Trabajo Descripcion'
        cols[13] = 'Pto. Trabajo'
        df_clean.columns = cols

    # Limpiar columnas de grupo y puesto de trabajo
    for col_name in ('Gr. Planif', 'Gr. planif', 'Gr.planif'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].astype(str).str.replace('CH01/', '', regex=False)
    for col_name in ('Pto. Trabajo', 'Pto. Trabajo Descripcion'):
        if col_name in df_clean.columns:
            df_clean[col_name] = df_clean[col_name].apply(clean_pto_trabajo)

    # Asignar Criterio y añadir los grupos calculados
    df_clean = df_clean.copy()
    df_clean['Criterio'] = criterios_col
    df_clean['Gr.planif'] = gr_planif_col
    df_clean['Gr.planif.PM'] = gr_planif_pm_col

    # Sobrescribir columnas R (17, Op.Ejec) y S (18, Op.Total) como número real,
    # no texto — de lo contrario la suma en la tabla dinámica no funciona.
    # Se reemplaza la columna completa (no .iloc in-place) porque el dtype
    # original suele ser 'string' (pandas StringDtype) y rechaza floats in-place.
    if 17 < len(df_clean.columns):
        while len(nuevos_ejec_col) < len(df_clean):
            nuevos_ejec_col.append(None)
        col_r = df_clean.columns[17]
        df_clean[col_r] = pd.to_numeric(pd.Series(nuevos_ejec_col), errors='coerce').values
    if 18 < len(df_clean.columns):
        while len(nuevos_totales_col) < len(df_clean):
            nuevos_totales_col.append(None)
        col_s = df_clean.columns[18]
        df_clean[col_s] = pd.to_numeric(pd.Series(nuevos_totales_col), errors='coerce').values
    if puestos_mapping and 'Pto. Trabajo Descripcion' in df_clean.columns and 'Pto. Trabajo' in df_clean.columns:
        df_clean['Pto. Trabajo Descripcion'] = df_clean['Pto. Trabajo'].apply(lambda x: str(puestos_mapping.get(x)).capitalize() if puestos_mapping.get(x) else 'N/A')

    if 'Pto. Trabajo' in df_clean.columns and 'Pto. Trabajo Descripcion' in df_clean.columns:
        cols = list(df_clean.columns)
        cols.remove('Pto. Trabajo')
        idx = cols.index('Pto. Trabajo Descripcion')
        cols.insert(idx + 1, 'Pto. Trabajo')
        df_clean = df_clean[cols]

    df_clean = df_clean.where(pd.notnull(df_clean), None)

    return df_clean, {
        'grupos': grupos,
        'total': {
            'cumple': total_ejec,
            'noCumple': total_totales - total_ejec,
            'total': total_totales,
            'cumplimiento': cumplimiento_total
        }
    }


# ──────────────────────────────────────────────────────────────────────────────
# PROCESADOR PRINCIPAL — MÚLTIPLES EXCELS CRUDOS
# ──────────────────────────────────────────────────────────────────────────────

def _build_tablas_static(wb, tablas_ws, stats, use_pto_trabajo):
    """Tablas resumen formateadas en hoja Tablas. Garantizado — siempre se crea."""
    from collections import OrderedDict

    if use_pto_trabajo:
        row_label = 'Pto. Trabajo'
        row_cols = ['ptoTrabajo', 'ptoTrabajoDesc']
    else:
        row_label = 'Gr. Planif'
        row_cols = ['grPlanif', 'grPlanifPM']

    current_row = 2

    def _add_table(title, grupos, value_cols_map):
        nonlocal current_row
        if not grupos:
            return
        data_keys = list(value_cols_map.values())
        display_names = list(value_cols_map.keys())
        num_cols = 1 + len(display_names)

        tablas_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        c = tablas_ws.cell(row=current_row, column=1, value=title)
        c.font = openpyxl.styles.Font(bold=True, size=13, color="1F4E79")
        c.fill = openpyxl.styles.PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        current_row += 1

        for ci, h in enumerate([row_label] + display_names, 1):
            c = tablas_ws.cell(row=current_row, column=ci, value=h)
            c.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
            c.fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        current_row += 1

        agrupado = OrderedDict()
        for g in grupos:
            key = g.get(row_cols[0], 'N/A')
            label = g.get(row_cols[1], '')
            display = f"{key} - {label}" if label and label not in ('N/A', key) else str(key)
            if key not in agrupado:
                agrupado[key] = {'display': display, 'vals': {dk: 0.0 for dk in data_keys}}
            for dk in data_keys:
                agrupado[key]['vals'][dk] += float(g.get(dk, 0) or 0)

        for info in agrupado.values():
            tablas_ws.cell(row=current_row, column=1, value=info['display']).font = openpyxl.styles.Font(size=11)
            for ci, dk in enumerate(data_keys, 2):
                c = tablas_ws.cell(row=current_row, column=ci, value=round(info['vals'][dk], 1))
                c.font = openpyxl.styles.Font(size=11)
                c.number_format = '#,##0.0'
            current_row += 1

        tablas_ws.cell(row=current_row, column=1, value="TOTAL").font = openpyxl.styles.Font(bold=True, size=11)
        for ci, dk in enumerate(data_keys, 2):
            total = sum(info['vals'][dk] for info in agrupado.values())
            c = tablas_ws.cell(row=current_row, column=ci, value=round(total, 1))
            c.font = openpyxl.styles.Font(bold=True, size=11)
            c.number_format = '#,##0.0'
        for ci in range(1, num_cols + 1):
            tablas_ws.cell(row=current_row, column=ci).fill = openpyxl.styles.PatternFill(
                start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        current_row += 2

    # Orden: Avisos, Ordenes, Trabajo Planificado, Programa Semanal, Plan Matriz
    _add_table('Avisos Pendientes', stats.get('avisos', {}).get('distribucion', []), OrderedDict([('Cantidad', 'cantidad')]))
    _add_table('Ordenes Pendientes', stats.get('ordenes', {}).get('distribucion', []), OrderedDict([('Cantidad', 'cantidad')]))
    _add_table('% Trabajo Planificado', stats.get('trabajo', {}).get('grupos', []), OrderedDict([
        ('Planificado', 'planificado'), ('Sin HR', 'sinHr'), ('Sin Horizonte', 'sinHorizonte'),
        ('Imprevistos', 'imprevistos'), ('Total HH', 'total')]))
    _add_table('Programa Semanal', stats.get('prog', {}).get('grupos', []), OrderedDict([
        ('Cumple', 'cumple'), ('No Cumple', 'noCumple'), ('Total Ops', 'total')]))
    _add_table('Plan Matriz', stats.get('plan', {}).get('grupos', []), OrderedDict([
        ('Cumple', 'cumple'), ('No Cumple', 'noCumple'), ('Total Ops', 'total')]))

    tablas_ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        tablas_ws.column_dimensions[col].width = 16


def _add_pivot_tables_com(output_path, use_pto_trabajo):
    """Abre XLSX con Excel COM. Crea PivotTables nativas en hoja 'Tablas'.
    Patrón: MonitoringApp (DispatchEx, constantes numéricas, Range objects, AddDataField)."""
    import os
    try:
        import pythoncom
        pythoncom.CoInitialize()
    except Exception:
        pass

    try:
        import win32com.client as win32
    except ImportError:
        print("[PivotTable] win32com no disponible.")
        return

    # Constantes numéricas (MonitoringApp pattern)
    XL_UP = -4162
    XL_TO_LEFT = -4159
    XL_DATABASE = 1
    XL_ROW_FIELD = 1
    XL_COL_FIELD = 2
    XL_SUM = -4157
    XL_COUNT = -4112
    XL_CALC_MANUAL = -4135
    XL_CALC_AUTO = -4105
    XL_TABULAR_ROW = 1
    XL_REPEAT_LABELS = 1

    def _campo_existe(pt, nombre):
        try:
            pt.PivotFields(nombre)
            return True
        except Exception:
            return False

    def _aplicar_formato(pt, nombre):
        """Aplica estilo naranja claro, quita subtotales y configura layout tabular."""
        try:
            pt.TableStyle2 = "PivotStyleMedium7"  # Naranja claro
        except Exception:
            try:
                pt.TableStyle2 = "PivotStyleMedium2"
            except Exception:
                pass
        try:
            pt.RowAxisLayout(1)  # Tabular
        except Exception:
            pass
        try:
            pt.RepeatAllLabels(1)
        except Exception:
            pass
        # Quitar subtotales de todos los row fields
        try:
            for pf in pt.PivotFields():
                try:
                    pf.Subtotals = (False,) * 12
                except Exception:
                    pass
        except Exception:
            pass
        # Quitar Grand Total de filas
        try:
            pt.RowGrand = True   # Mantener total general
        except Exception:
            pass

    excel = None
    wb = None
    excel_pid = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        try:
            import win32process
            excel_pid = win32process.GetWindowThreadProcessId(excel.Hwnd)[1]
        except Exception:
            excel_pid = None

        abs_path = os.path.abspath(output_path)
        if not os.path.exists(abs_path):
            print(f"[PivotTable] Archivo {abs_path} no encontrado.")
            return

        wb = excel.Workbooks.Open(abs_path)
        
        # Set calculation AFTER opening workbook, otherwise Excel COM throws error
        try:
            excel.Calculation = XL_CALC_MANUAL
        except:
            pass
        
        sheet_names = [s.Name for s in wb.Worksheets]

        # Recrear hoja 'Tablas'
        if "Tablas" in sheet_names:
            tablas_ws = wb.Worksheets("Tablas")
            tablas_ws.Cells.Clear()
        else:
            tablas_ws = wb.Worksheets.Add(Before=wb.Worksheets(1))
            tablas_ws.Name = "Tablas"

        def _first_row_with_data(ws, col=1):
            for r in range(1, 20):
                v = ws.Cells(r, col).Value
                if v is not None and str(v).strip() not in ('', 'nan'):
                    return r
            return 1

        def _find_header_row(ws):
            """Encontrar fila que parece header (sin números, texto significativo)."""
            for r in range(1, 8):
                cell_text = []
                for c in range(1, min(20, ws.UsedRange.Columns.Count + 1)):
                    v = ws.Cells(r, c).Value
                    if v is not None:
                        cell_text.append(str(v).strip())
                txt = ' '.join(cell_text).lower()
                # Si esta fila contiene palabras clave tipo header
                if any(k in txt for k in ('proceso', 'gr.', 'planif', 'orden', 'aviso', 'area', 'fase', 'ano natural')):
                    return r
            return 1

        current_row = 1

        # Estructura exacta replicando el Excel de referencia KPI GSYS SEM20:
        # (keyword_sheet, titulo, col_field_kw, [(data_kw, agg, etiqueta)])
        pivot_defs = [
            ('Avisos',              'Avisos Pendientes',      None,       [('número de avisos pendientes', XL_COUNT, 'Cantidad')]),
            ('rdenes',              'Órdenes Pendientes',      None,       [('número de órdenes', XL_COUNT, 'Cantidad')]),
            ('Trabajo Planificado', '% Trabajo Planificado', 'criterio', [('hh totales reales', XL_SUM, 'Suma HH Totales Reales', '#,##0')]),
            ('Programa Semanal',    'Programa Semanal',      'criterio', [('total op. programadas', XL_SUM, 'Suma Total Op. Programadas')]),
            ('Plan Matriz',         'Plan Matriz',           'criterio', [('cantidad de operaciones totales', XL_SUM, 'Suma Cantidad Ops')]),
        ]

        for sheet_keyword, titulo, col_field_kw, data_defs in pivot_defs:
            src_ws = None
            for sn in sheet_names:
                if sheet_keyword.lower() in str(sn).lower():
                    src_ws = wb.Worksheets(sn)
                    break
            if src_ws is None:
                continue

            # Detectar rango de datos (xlUp / xlToLeft — MonitoringApp pattern)
            ult_f = src_ws.Cells(src_ws.Rows.Count, 1).End(XL_UP).Row
            ult_c = src_ws.Cells(1, src_ws.Columns.Count).End(XL_TO_LEFT).Column
            if ult_f < 2:
                continue

            rng = src_ws.Range(src_ws.Cells(1, 1), src_ws.Cells(ult_f, ult_c))

            # PivotCache (MonitoringApp pattern)
            try:
                pc = wb.PivotCaches().Create(XL_DATABASE, rng)
            except Exception as e:
                print(f"[PivotTable] Error CreatePivotCache para {sheet_keyword}: {e}")
                continue

            # Limpiar pivot anterior
            nombre_pt = f"PT_{sheet_keyword.replace(' ','_')[:20]}"
            try:
                tablas_ws.PivotTables(nombre_pt).TableRange2.Clear()
            except Exception:
                pass

            # --- Título sobre la pivot ---
            if current_row > 1:
                current_row += 1  # fila extra de separación
            tablas_ws.Cells(current_row, 1).Value = titulo
            try:
                tablas_ws.Cells(current_row, 1).Font.Bold = True
                tablas_ws.Cells(current_row, 1).Font.Size = 13
                tablas_ws.Cells(current_row, 1).Font.Color = 0x000000  # Negro
            except Exception:
                pass
            pivot_row = current_row + 1

            # Crear PivotTable
            try:
                pt = pc.CreatePivotTable(tablas_ws.Cells(pivot_row, 1), nombre_pt)
            except Exception as e:
                print(f"[PivotTable] Error CreatePivotTable para {sheet_keyword}: {e}")
                continue

            # --- Row fields: Proceso (pos1) -> Gr.planif (pos2) -> Gr.planif.PM (pos3) ---
            header_row = _find_header_row(src_ws)

            def _set_row_field(pt, col_idx, position):
                """Configura una columna como Row Field en la posición dada."""
                if col_idx and _campo_existe(pt, col_idx):
                    try:
                        pf = pt.PivotFields(col_idx)
                        pf.Orientation = XL_ROW_FIELD
                        pf.Position = position
                        return True
                    except Exception:
                        pass
                return False

            def _find_col_idx(kw_list, exclude_kw=None):
                """Busca índice de columna por lista de keywords."""
                for kw in kw_list:
                    for c in range(1, min(ult_c + 1, 40)):
                        val = str(src_ws.Cells(header_row, c).Value or '').strip().lower()
                        if exclude_kw and any(ex in val for ex in exclude_kw):
                            continue
                        if kw in val:
                            return c
                return None

            # Decidir si row fields son Gr.Planif o Pto. Trabajo según use_pto_trabajo
            if use_pto_trabajo:
                col2_kw   = ['pto. trabajo']
                col2_ex   = ['descripcion', 'desc']
                col3_kw   = ['pto. trabajo descripcion', 'puesto de trabajo']
                col3_ex   = []
            else:
                col2_kw   = ['gr. planif', 'gr.planif']
                col2_ex   = ['pm', 'pm_1']
                col3_kw   = ['gr.planif.pm', 'gr. planif.pm']
                col3_ex   = ['_1']

            proceso_idx   = _find_col_idx(['proceso'], exclude_kw=['subprocess', 'subproceso'])
            col2_idx      = _find_col_idx(col2_kw, exclude_kw=col2_ex)
            col3_idx      = _find_col_idx(col3_kw, exclude_kw=col3_ex)

            pos = 1
            if _set_row_field(pt, proceso_idx, pos): pos += 1
            if _set_row_field(pt, col2_idx, pos):    pos += 1
            if _set_row_field(pt, col3_idx, pos):    pos += 1

            # --- Column field: Criterio (si aplica) ---
            if col_field_kw:
                crit_idx = _find_col_idx([col_field_kw])
                if crit_idx and _campo_existe(pt, crit_idx):
                    try:
                        pt.PivotFields(crit_idx).Orientation = XL_COL_FIELD
                    except Exception:
                        pass

            # --- Data fields ---
            found_data = False
            for data_def in data_defs:
                dh_kw, dh_agg, dh_label = data_def[0], data_def[1], data_def[2]
                dh_fmt = data_def[3] if len(data_def) > 3 else None
                col_idx = _find_col_idx([dh_kw])
                if col_idx is None:
                    continue
                if _campo_existe(pt, col_idx):
                    try:
                        pf_data = pt.AddDataField(pt.PivotFields(col_idx), dh_label, dh_agg)
                        if dh_fmt:
                            pf_data.NumberFormat = dh_fmt
                        found_data = True
                    except Exception:
                        pass
            # Fallback: contar por una columna que NO esté ya usada como row/col field.
            # AddDataField sobre un campo que ya es RowField lo reconvierte a DataField
            # y lo saca de las filas (por eso NO se puede reusar col2_idx/col3_idx aquí).
            if not found_data:
                ocupadas = {proceso_idx, col2_idx, col3_idx}
                if col_field_kw:
                    ocupadas.add(_find_col_idx([col_field_kw]))
                fallback_idx = next((c for c in range(1, min(ult_c + 1, 40)) if c not in ocupadas), None)
                if fallback_idx and _campo_existe(pt, fallback_idx):
                    try:
                        pt.AddDataField(pt.PivotFields(fallback_idx), 'Cantidad', XL_COUNT)
                    except Exception:
                        pass

            _aplicar_formato(pt, nombre_pt)

            # Avanzar: 4 filas después de la tabla
            current_row = pt.TableRange2.Row + pt.TableRange2.Rows.Count + 4


        tablas_ws.Columns(1).ColumnWidth = 40
        excel.Calculation = XL_CALC_AUTO
        wb.Save()
        print("[PivotTable] OK - PivotTables nativas creadas en hoja Tablas")

    except Exception as e:
        print(f"[PivotTable] Error: {e}")
    finally:
        # Cleanup (MonitoringApp pattern)
        try:
            if excel:
                excel.ScreenUpdating = True
                excel.DisplayAlerts = True
                excel.Calculation = XL_CALC_AUTO
                excel.EnableEvents = True
        except Exception:
            pass
        if wb:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        # Liberar referencias COM explícitamente antes de forzar el recolector
        # cíclico: pywin32 no libera los objetos COM solo con Quit() si Python
        # aún los referencia (pt, pc, rng, src_ws, tablas_ws siguen vivos en
        # este frame hasta que el finally termina), dejando EXCEL.EXE huérfano
        # en segundo plano y bloqueando el archivo generado.
        wb = None
        excel = None
        pt = None
        pc = None
        rng = None
        src_ws = None
        tablas_ws = None
        try:
            import gc
            gc.collect()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        # Red de seguridad: si pese al cleanup anterior EXCEL.EXE sigue vivo
        # (referencia COM oculta en algún helper), forzar su cierre por PID
        # para no dejar el archivo bloqueado.
        if excel_pid:
            try:
                import time as _time, subprocess as _subprocess, win32process as _win32process, win32con as _win32con
                _time.sleep(1.0)
                handle = _win32process.OpenProcess(_win32con.PROCESS_QUERY_INFORMATION, False, excel_pid)
                still_alive = _win32process.GetExitCodeProcess(handle) == _win32con.STILL_ACTIVE
                if still_alive:
                    _subprocess.run(["taskkill", "/F", "/PID", str(excel_pid)], capture_output=True)
            except Exception:
                pass

def process_kpi_excels(file_paths, semana_num, output_path,
                       ots_mapping=None, export_ops_mapping=None,
                       puestos_mapping=None, grupos_mapping=None,
                       use_pto_trabajo=False, metadata=None):
    """
    Procesa los 5 archivos SAP crudos (flujo Múltiples Excel), aplica la lógica de
    preprocesamiento y genera el Excel consolidado.
    - ots_mapping: dict {orden: (gr_planif_code, gr_planif_pm)} para Trabajo Planificado.
    - export_ops_mapping: dict {orden: cant_ops} para corregir totales en Plan Matriz.
    - grupos_mapping: dict {codigo_grupo: nombre_grupo} desde DB o PLANNING_GROUP_MAP.
    No afecta a process_ready_excel (flujo Excel Consolidado Listo).
    """
    if ots_mapping is None:
        ots_mapping = {}
    if export_ops_mapping is None:
        export_ops_mapping = {}
    if grupos_mapping is None:
        grupos_mapping = dict(PLANNING_GROUP_MAP)

    print(f"[ExcelProcessor] Procesando archivos múltiples semana {semana_num}...")

    # Extraer y preprocesar cada archivo (archivos opcionales — usar .get() con fallback)
    empty_df = pd.DataFrame()
    empty_av  = {'total': 0, 'distribucion': []}
    empty_tp  = {'grupos': [], 'total': {'planificado': 0, 'sinHr': 0, 'sinHorizonte': 0, 'imprevistos': 0, 'total': 0, 'cumplimiento': 0}}
    empty_ps  = {'grupos': [], 'total': {'cumple': 0, 'noCumple': 0, 'total': 0, 'cumplimiento': 0}}

    def _safe(key, fn, *args, **kwargs):
        path = file_paths.get(key)
        if path and os.path.exists(path):
            try:
                return fn(path, *args, **kwargs)
            except Exception as e:
                print(f"[ExcelProcessor] Error extrayendo {key}: {e}")
        if key in ('avisos', 'ordenes'):
            return empty_df, dict(empty_av)
        if key == 'trabajoPlanificado':
            return empty_df, dict(empty_tp)
        return empty_df, dict(empty_ps)

    df_avisos,  stats_avisos  = _safe('avisos',            extract_avisos, puestos_mapping=puestos_mapping, grupos_mapping=grupos_mapping)
    df_ordenes, stats_ordenes = _safe('ordenes',           extract_ordenes, puestos_mapping=puestos_mapping, grupos_mapping=grupos_mapping)
    df_trabajo, stats_trabajo = _safe('trabajoPlanificado', extract_trabajo_planificado, ots_mapping, puestos_mapping, grupos_mapping=grupos_mapping)
    df_prog,    stats_prog    = _safe('programaSemanal',   extract_programa_semanal, puestos_mapping=puestos_mapping, grupos_mapping=grupos_mapping)
    df_plan,    stats_plan    = _safe('planMatriz',        extract_plan_matriz, export_ops_mapping, puestos_mapping, grupos_mapping=grupos_mapping)

    # Generar XLSX consolidado
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Hoja 1: Tablas (estáticas formateadas — garantizadas) ──
    tablas_ws = wb.create_sheet('Tablas', 0)
    combined_stats = {
        'trabajo': stats_trabajo, 'prog': stats_prog, 'plan': stats_plan,
        'avisos': stats_avisos, 'ordenes': stats_ordenes,
    }
    try:
        _build_tablas_static(wb, tablas_ws, combined_stats, use_pto_trabajo)
    except Exception as e:
        print(f"[ExcelProcessor] Error hoja Tablas: {e}")

    # ── Hojas de detalle ──
    sheet_pairs = [
        ('Avisos Pendientes',    df_avisos),
        ('Órdenes Pendientes',   df_ordenes),
        ('% Trabajo Planificado', df_trabajo),
        ('Programa Semanal',     df_prog),
        ('Plan Matriz',          df_plan),
    ]

    for sheet_name, df in sheet_pairs:
        ws = wb.create_sheet(sheet_name)
        if df is not None and len(df) > 0:
            try:
                # Filtrar filas de subtotal intermedio ('Resultado'), conservar 'Resultado total'
                df_clean = filtrar_filas_resultado_intermedias(df)

                for r_idx, row in enumerate(dataframe_to_rows(df_clean, index=False, header=True)):
                    for c_idx, val in enumerate(row):
                        ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)

                # Aplicar formato tabla a la hoja
                aplicar_formato_tabla_openpyxl(ws, sheet_name)

            except Exception as e:
                print(f"[ExcelProcessor] Error escribiendo hoja {sheet_name}: {e}")

    wb.save(output_path)
    print(f"[ExcelProcessor] XLSX consolidado guardado en: {output_path}")

    # Crear PivotTables nativas con Excel COM (drill-down real)
    _add_pivot_tables_com(output_path, use_pto_trabajo)

    # Construir respuesta de KPIs en formato estándar
    tp   = stats_trabajo['total']
    prog = stats_prog['total']
    plan = stats_plan['total']

    return {
        'semana': semana_num,
        'indicadores': {
            'avisosPendientes':  stats_avisos['total'],
            'ordenesPendientes': stats_ordenes['total'],
            'trabajoPlanificado': int(round(tp['cumplimiento']   * 100)),
            'programaSemanal':   int(round(prog['cumplimiento']  * 100)),
            'planMatriz':        int(round(plan['cumplimiento']   * 100))
        },
        'resumenAvisos': {
            'total': stats_avisos['total'],
            'distribucion': stats_avisos['distribucion']
        },
        'resumenOrdenes': {
            'total': stats_ordenes['total'],
            'distribucion': stats_ordenes['distribucion']
        },
        'trabajoPlanificado': {
            'grupos': stats_trabajo['grupos'],
            'total':  tp
        },
        'programaSemanal': {
            'grupos': stats_prog['grupos'],
            'total':  prog
        },
        'planMatriz': {
            'grupos': stats_plan['grupos'],
            'total':  plan
        }
    }


# ──────────────────────────────────────────────────────────────────────────────
# FUNCIÓN DE PREVISUALIZACIÓN RÁPIDA
# ──────────────────────────────────────────────────────────────────────────────

def preview_file(path, file_type):
    """
    Hace un análisis rápido de un archivo SAP crudo para retroalimentación inmediata.
    Retorna un dict con conteo de filas y mensaje descriptivo.
    """
    try:
        df = read_raw_sap_file(path)

        # Determinar offset de encabezado según tipo
        double_header_types = ('trabajoPlanificado', 'programaSemanal', 'planMatriz')
        header_offset = 2 if file_type in double_header_types else 1

        data_rows = df.iloc[header_offset:-1] if len(df) > header_offset + 1 else df.iloc[header_offset:]
        valid_rows = [row for _, row in data_rows.iterrows() if not is_resultado_row(list(row))]
        count = len(valid_rows)

        labels = {
            'avisos':            f'{count} avisos pendientes detectados',
            'ordenes':           f'{count} órdenes pendientes detectadas',
            'trabajoPlanificado': f'{count} operaciones de trabajo detectadas',
            'programaSemanal':   f'{count} operaciones de programa detectadas',
            'planMatriz':        f'{count} operaciones de plan matriz detectadas',
        }
        message = labels.get(file_type, f'{count} filas detectadas')
        return {'success': True, 'rows': count, 'message': message}
    except Exception as e:
        return {'success': False, 'error': str(e), 'rows': 0}


# ──────────────────────────────────────────────────────────────────────────────
# PROCESADOR DE EXCEL CONSOLIDADO (modo "Listo")
# ──────────────────────────────────────────────────────────────────────────────

def _find_col_idx(columns, *keywords):
    """Busca el índice de columna por nombre (insensible a mayúsculas y parcial)."""
    # 1. Búsqueda exacta
    for i, col in enumerate(columns):
        col_str = str(col).strip().lower()
        for kw in keywords:
            if kw.lower() == col_str:
                return i
                
    # 2. Búsqueda parcial con prevención de colisiones
    for i, col in enumerate(columns):
        col_str = str(col).strip().lower()
        for kw in keywords:
            # Evitar que 'gr. planif' haga match con 'gr. planif.pm'
            if 'planif' in kw.lower() and 'pm' not in kw.lower() and 'pm' in col_str:
                continue
            # Evitar que 'pto. trabajo' haga match con 'pto. trabajo descripcion'
            if 'trabajo' in kw.lower() and 'desc' not in kw.lower() and 'desc' in col_str:
                continue
            if kw.lower() in col_str:
                return i
    return -1


def _filter_data_rows(df):
    """Elimina filas de total/subtotal SAP de un DataFrame con header."""
    def is_bad_row(row):
        for v in list(row)[:8]:
            if str(v).strip() in ('Resultado total', 'Resultado'):
                return True
        return False
    mask = ~df.apply(is_bad_row, axis=1)
    return df[mask]


def _read_sheet_smart(path, sheet_name):
    """
    Lee una hoja del XLSX intentando detectar automáticamente la fila de encabezado.
    Retorna (df_con_headers, offset_info).
    """
    try:
        df_raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    except Exception:
        return None

    if df_raw is None or len(df_raw) == 0:
        return None

    # Intentar con header en fila 0
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=0)
        return df
    except Exception:
        return None


def process_ready_excel(file_path, semana_num, puestos_mapping=None):
    """
    Procesa un Excel consolidado KPI (generado por process_kpi_excels o macro VBA).
    Lee hoja por hoja usando detección de columnas por nombre.
    - puestos_mapping: dict {puesto_trabajo: descripcion} para enriquecer distribución/grupos.
    """
    print(f"[ExcelProcessor] Procesando Excel consolidado semana {semana_num}...")

    try:
        xl = pd.ExcelFile(file_path)
        sheet_names = xl.sheet_names
    except Exception as e:
        raise ValueError(f"No se pudo abrir el archivo: {e}")

    # ── Avisos Pendientes ──────────────────────────────────────────────────────
    total_avisos_count = 0
    group_avisos = {}

    avi_sheet = next((s for s in sheet_names if 'aviso' in s.lower()), None)
    if avi_sheet:
        df_avi = _read_sheet_smart(file_path, avi_sheet)
        if df_avi is not None:
            df_avi = _filter_data_rows(df_avi)
            col_proc = _find_col_idx(df_avi.columns, 'proceso')
            col_gp   = _find_col_idx(df_avi.columns, 'gr. planif', 'gr.planif', 'grplanif')
            col_gppm = _find_col_idx(df_avi.columns, 'gr.planif.pm', 'gr. planif.pm')
            col_pto  = _find_col_idx(df_avi.columns, 'pto. trabajo', 'pto trabajo', 'puesto de trabajo', 'pto trabajo')

            for _, row in df_avi.iterrows():
                row_l = list(row)
                if is_resultado_row(row_l):
                    continue

                proceso_raw  = str(row_l[col_proc]  if col_proc >= 0  else row_l[2]).strip()
                gr_planif    = str(row_l[col_gp]    if col_gp >= 0    else row_l[11]).strip()
                gr_planif_pm = str(row_l[col_gppm]  if col_gppm >= 0  else row_l[10]).strip()
                pto_trabajo  = clean_pto_trabajo(str(row_l[col_pto] if col_pto >= 0 else 'N/A').strip())

                if not proceso_raw or proceso_raw in ('nan', ''):
                    continue

                proceso = get_safe_proceso(proceso_raw)
                gr_planif = strip_ch01(gr_planif)
                if not gr_planif or gr_planif == 'nan':
                    gr_planif = 'N/A'
                if not gr_planif_pm or gr_planif_pm == 'nan':
                    gr_planif_pm = PLANNING_GROUP_MAP.get(gr_planif, gr_planif)

                key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
                group_avisos[key] = group_avisos.get(key, 0) + 1
                total_avisos_count += 1

    # ── Órdenes Pendientes ────────────────────────────────────────────────────
    total_ordenes_count = 0
    group_ordenes = {}

    ord_sheet = next((s for s in sheet_names if 'orden' in s.lower() or 'órden' in s.lower() or 'ot' in s.lower()), None)
    if ord_sheet:
        df_ord = _read_sheet_smart(file_path, ord_sheet)
        if df_ord is not None:
            df_ord = _filter_data_rows(df_ord)
            col_proc = _find_col_idx(df_ord.columns, 'proceso')
            col_gp   = _find_col_idx(df_ord.columns, 'gr. planif', 'gr.planif')
            col_gppm = _find_col_idx(df_ord.columns, 'gr.planif.pm', 'gr. planif.pm')
            col_pto  = _find_col_idx(df_ord.columns, 'pto. trabajo', 'pto trabajo', 'puesto de trabajo')

            for _, row in df_ord.iterrows():
                row_l = list(row)
                if is_resultado_row(row_l):
                    continue

                proceso_raw  = str(row_l[col_proc]  if col_proc >= 0  else row_l[3]).strip()
                gr_planif    = str(row_l[col_gp]    if col_gp >= 0    else row_l[1]).strip()
                gr_planif_pm = str(row_l[col_gppm]  if col_gppm >= 0  else row_l[0]).strip()
                pto_trabajo  = clean_pto_trabajo(str(row_l[col_pto] if col_pto >= 0 else 'N/A').strip())

                if not proceso_raw or proceso_raw in ('nan', ''):
                    continue

                proceso = get_safe_proceso(proceso_raw)
                gr_planif = strip_ch01(gr_planif)
                if not gr_planif or gr_planif == 'nan':
                    gr_planif = 'N/A'
                if not gr_planif_pm or gr_planif_pm == 'nan':
                    gr_planif_pm = PLANNING_GROUP_MAP.get(gr_planif, gr_planif)

                key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
                group_ordenes[key] = group_ordenes.get(key, 0) + 1
                total_ordenes_count += 1

    # ── % Trabajo Planificado ─────────────────────────────────────────────────
    tp_sheet = next((s for s in sheet_names if 'trabajo' in s.lower() or 'planificad' in s.lower()), None)
    tp_rows_raw = []
    if tp_sheet:
        tp_rows_raw = pd.read_excel(file_path, sheet_name=tp_sheet, header=None).values.tolist()

    total_tp_cumplimiento_value = extract_total_percentage(tp_rows_raw, 2)

    group_tp            = {}
    total_tp_planificado = 0.0
    total_tp_sin_hr     = 0.0
    total_tp_imprevistos = 0.0

    if tp_sheet:
        df_tp = _read_sheet_smart(file_path, tp_sheet)
        if df_tp is not None:
            df_tp = _filter_data_rows(df_tp)
            col_proc    = _find_col_idx(df_tp.columns, 'proceso')
            col_gp      = _find_col_idx(df_tp.columns, 'gr. planif', 'gr.planif')
            col_gppm    = _find_col_idx(df_tp.columns, 'gr.planif.pm', 'gr. planif.pm')
            col_criterio= _find_col_idx(df_tp.columns, 'criterio')
            col_hh_real = _find_col_idx(df_tp.columns, 'hh totales reales', 'hh total', 'trabajo real', 'trab.real', 'trabajo tot')
            col_pto     = _find_col_idx(df_tp.columns, 'pto. trabajo', 'pto trabajo', 'puesto de trabajo')

            for _, row in df_tp.iterrows():
                row_l = list(row)
                if is_resultado_row(row_l):
                    continue

                proceso_raw  = str(row_l[col_proc]  if col_proc >= 0  else row_l[2]).strip()
                gr_planif    = str(row_l[col_gp]    if col_gp >= 0    else '').strip()
                gr_planif_pm = str(row_l[col_gppm]  if col_gppm >= 0  else '').strip()
                criterio_raw = str(row_l[col_criterio] if col_criterio >= 0 else 'Planificado').strip().lower()
                
                # In process_ready_excel, the numbers are already standard floats, so we just get_num
                hh_real      = get_num(row_l, col_hh_real if col_hh_real >= 0 else -2)
                pto_trabajo  = clean_pto_trabajo(str(row_l[col_pto] if col_pto >= 0 else 'N/A').strip())

                if not proceso_raw or proceso_raw in ('nan', ''):
                    continue

                proceso = get_safe_proceso(proceso_raw)
                gr_planif = strip_ch01(gr_planif) or 'N/A'
                if not gr_planif_pm or gr_planif_pm == 'nan':
                    gr_planif_pm = PLANNING_GROUP_MAP.get(gr_planif, gr_planif)

                if is_total_or_invalid_row(proceso, gr_planif, gr_planif_pm):
                    continue

                key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
                if key not in group_tp:
                    group_tp[key] = {'planificado': 0.0, 'sinHr': 0.0, 'imprevistos': 0.0}

                if 'planificad' in criterio_raw:
                    group_tp[key]['planificado'] += hh_real
                    total_tp_planificado += hh_real
                elif 'sin hr' in criterio_raw or 'sin horizonte' in criterio_raw:
                    group_tp[key]['sinHr'] += hh_real
                    total_tp_sin_hr += hh_real
                elif 'imprevist' in criterio_raw:
                    group_tp[key]['imprevistos'] += hh_real
                    total_tp_imprevistos += hh_real

    total_tp_total       = total_tp_planificado + total_tp_sin_hr + total_tp_imprevistos
    total_tp_cumplimiento = total_tp_planificado / total_tp_total if total_tp_total > 0 else 0.0
    cump_trabajo_planificado = []
    for key in sorted(group_tp.keys()):
        proceso, gp, gppm, pto = key.split('||')
        pto_desc = str(puestos_mapping.get(pto, 'N/A')).capitalize() if puestos_mapping and puestos_mapping.get(pto) else 'N/A'
        vals = group_tp[key]
        rt   = vals['planificado'] + vals['sinHr'] + vals['imprevistos']
        c    = vals['planificado'] / rt if rt > 0 else 0.0
        cump_trabajo_planificado.append({
            'proceso': proceso, 'grPlanif': gp, 'grPlanifPM': gppm, 'ptoTrabajo': pto, 'ptoTrabajoDesc': pto_desc,
            'planificado': vals['planificado'], 'sinHr': vals['sinHr'],
            'imprevistos': vals['imprevistos'], 'total': rt, 'cumplimiento': c
        })

    # ── Programa Semanal ──────────────────────────────────────────────────────
    prog_sheet  = next((s for s in sheet_names if 'programa' in s.lower() or 'semanal' in s.lower()), None)
    prog_rows_raw = []
    if prog_sheet:
        prog_rows_raw = pd.read_excel(file_path, sheet_name=prog_sheet, header=None).values.tolist()

    total_prog_cumplimiento_value = extract_total_percentage(prog_rows_raw, 4)

    group_prog      = {}
    total_prog_cumple    = 0.0
    total_prog_no_cumple = 0.0
    total_prog_total_ops = 0.0
    total_prog_ind_cumple= 0.0

    if prog_sheet:
        df_prog = _read_sheet_smart(file_path, prog_sheet)
        if df_prog is not None:
            df_prog = _filter_data_rows(df_prog)
            col_proc    = _find_col_idx(df_prog.columns, 'proceso')
            col_gp      = _find_col_idx(df_prog.columns, 'gr. planif', 'gr.planif')
            col_gppm    = _find_col_idx(df_prog.columns, 'gr.planif.pm', 'gr. planif.pm')
            col_criterio= _find_col_idx(df_prog.columns, 'criterio')
            col_ind     = _find_col_idx(df_prog.columns, 'indicador cumple', 'ind cumple')
            col_total   = _find_col_idx(df_prog.columns, 'total op', 'total op. programadas')
            col_pto     = _find_col_idx(df_prog.columns, 'pto. trabajo', 'pto trabajo', 'puesto de trabajo')

            for _, row in df_prog.iterrows():
                row_l = list(row)
                if is_resultado_row(row_l):
                    continue

                proceso_raw  = str(row_l[col_proc]  if col_proc >= 0  else row_l[4]).strip()
                gr_planif    = str(row_l[col_gp]    if col_gp >= 0    else row_l[1]).strip()
                gr_planif_pm = str(row_l[col_gppm]  if col_gppm >= 0  else row_l[0]).strip()
                criterio_raw = str(row_l[col_criterio] if col_criterio >= 0 else 'Cumple').strip().lower()
                ind_cumple   = get_num(row_l, col_ind   if col_ind >= 0   else 15)
                total_op     = get_num(row_l, col_total if col_total >= 0 else 16)
                pto_trabajo  = clean_pto_trabajo(str(row_l[col_pto] if col_pto >= 0 else 'N/A').strip())

                if not proceso_raw or proceso_raw in ('nan', ''):
                    continue

                proceso = get_safe_proceso(proceso_raw)
                gr_planif = strip_ch01(gr_planif) or 'N/A'
                if not gr_planif_pm or gr_planif_pm == 'nan':
                    gr_planif_pm = PLANNING_GROUP_MAP.get(gr_planif, gr_planif)

                if is_total_or_invalid_row(proceso, gr_planif, gr_planif_pm):
                    continue

                key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
                if key not in group_prog:
                    group_prog[key] = {'cumple': 0.0, 'noCumple': 0.0,
                                       'sumTotalOp': 0.0, 'sumIndicadorCumple': 0.0}

                if 'cumple' in criterio_raw and 'no cumple' not in criterio_raw:
                    group_prog[key]['cumple'] += total_op
                    total_prog_cumple += total_op
                else:
                    group_prog[key]['noCumple'] += total_op
                    total_prog_no_cumple += total_op

                group_prog[key]['sumTotalOp']         += total_op
                group_prog[key]['sumIndicadorCumple'] += ind_cumple
                total_prog_total_ops  += total_op
                total_prog_ind_cumple += ind_cumple

    total_prog_cumplimiento = (total_prog_ind_cumple / total_prog_total_ops
                               if total_prog_total_ops > 0 else 0.0)
    cump_programa_semanal = []
    for key in sorted(group_prog.keys()):
        proceso, gp, gppm, pto = key.split('||')
        pto_desc = str(puestos_mapping.get(pto, 'N/A')).capitalize() if puestos_mapping and puestos_mapping.get(pto) else 'N/A'
        vals = group_prog[key]
        c = vals['sumIndicadorCumple'] / vals['sumTotalOp'] if vals['sumTotalOp'] > 0 else 0.0
        cump_programa_semanal.append({
            'proceso': proceso, 'grPlanif': gp, 'grPlanifPM': gppm, 'ptoTrabajo': pto, 'ptoTrabajoDesc': pto_desc,
            'cumple': vals['cumple'], 'noCumple': vals['noCumple'],
            'total': vals['sumTotalOp'], 'cumplimiento': c
        })

    # ── Plan Matriz ───────────────────────────────────────────────────────────
    mtz_sheet   = next((s for s in sheet_names if 'matriz' in s.lower()), None)
    mtz_rows_raw = []
    if mtz_sheet:
        mtz_rows_raw = pd.read_excel(file_path, sheet_name=mtz_sheet, header=None).values.tolist()

    group_matriz     = {}
    total_mtz_ejec   = 0.0
    total_mtz_totales = 0.0

    if mtz_sheet:
        df_mtz = _read_sheet_smart(file_path, mtz_sheet)
        if df_mtz is not None:
            df_mtz = _filter_data_rows(df_mtz)
            col_proc    = _find_col_idx(df_mtz.columns, 'proceso')
            col_gp      = _find_col_idx(df_mtz.columns, 'gr. planif', 'gr.planif')
            col_gppm    = _find_col_idx(df_mtz.columns, 'gr.planif.pm', 'gr. planif.pm')
            col_criterio= _find_col_idx(df_mtz.columns, 'criterio')
            col_ejec    = _find_col_idx(df_mtz.columns, 'op. ejec', 'operaciones ejec', 'cantidad de operaciones ejec')
            col_total   = _find_col_idx(df_mtz.columns, 'op. tot', 'operaciones tot', 'cantidad de operaciones tot')
            col_pto     = _find_col_idx(df_mtz.columns, 'pto. trabajo', 'pto trabajo', 'puesto de trabajo')

            for _, row in df_mtz.iterrows():
                row_l = list(row)
                if is_resultado_row(row_l):
                    continue

                proceso_raw  = str(row_l[col_proc]  if col_proc >= 0  else row_l[2]).strip()
                gr_planif    = str(row_l[col_gp]    if col_gp >= 0    else row_l[5]).strip()
                gr_planif_pm = str(row_l[col_gppm]  if col_gppm >= 0  else row_l[4]).strip()
                criterio_raw = str(row_l[col_criterio] if col_criterio >= 0 else 'Cumple').strip().lower()
                op_ejec      = get_num(row_l, col_ejec  if col_ejec >= 0  else 17)
                op_total     = get_num(row_l, col_total if col_total >= 0 else 18)
                pto_trabajo  = clean_pto_trabajo(str(row_l[col_pto] if col_pto >= 0 else 'N/A').strip())

                if not proceso_raw or proceso_raw in ('nan', ''):
                    continue

                proceso = get_safe_proceso(proceso_raw)
                gr_planif = strip_ch01(gr_planif) or 'N/A'
                if not gr_planif_pm or gr_planif_pm == 'nan':
                    gr_planif_pm = PLANNING_GROUP_MAP.get(gr_planif, gr_planif)

                if is_total_or_invalid_row(proceso, gr_planif, gr_planif_pm):
                    continue

                key = f"{proceso}||{gr_planif}||{gr_planif_pm}||{pto_trabajo}"
                if key not in group_matriz:
                    group_matriz[key] = {'cumple': 0.0, 'noCumple': 0.0}

                if 'cumple' in criterio_raw and 'no cumple' not in criterio_raw:
                    group_matriz[key]['cumple'] += op_total
                    total_mtz_ejec += op_ejec
                else:
                    group_matriz[key]['noCumple'] += op_total

                total_mtz_totales += op_total

    total_mtz_cumplimiento = total_mtz_ejec / total_mtz_totales if total_mtz_totales > 0 else 0.0
    cump_plan_matriz = []
    for key in sorted(group_matriz.keys()):
        proceso, gp, gppm, pto = key.split('||')
        pto_desc = str(puestos_mapping.get(pto, 'N/A')).capitalize() if puestos_mapping and puestos_mapping.get(pto) else 'N/A'
        vals = group_matriz[key]
        rt   = vals['cumple'] + vals['noCumple']
        c    = vals['cumple'] / rt if rt > 0 else 0.0
        cump_plan_matriz.append({
            'proceso': proceso, 'grPlanif': gp, 'grPlanifPM': gppm, 'ptoTrabajo': pto, 'ptoTrabajoDesc': pto_desc,
            'cumple': vals['cumple'], 'noCumple': vals['noCumple'],
            'total': rt, 'cumplimiento': c
        })

    return {
        'semana': semana_num,
        'indicadores': {
            'avisosPendientes':  total_avisos_count,
            'ordenesPendientes': total_ordenes_count,
            'trabajoPlanificado': int(round((total_tp_cumplimiento_value
                                             if total_tp_cumplimiento_value is not None
                                             else total_tp_cumplimiento) * 100)),
            'programaSemanal':   int(round((total_prog_cumplimiento_value
                                            if total_prog_cumplimiento_value is not None
                                            else total_prog_cumplimiento) * 100)),
            'planMatriz':        int(round(total_mtz_cumplimiento * 100))
        },
        'resumenAvisos': {
            'total': total_avisos_count,
            'distribucion': [
                {'proceso': k.split('||')[0], 'grPlanif': k.split('||')[1],
                 'grPlanifPM': k.split('||')[2],
                 'ptoTrabajo': k.split('||')[3] if len(k.split('||')) > 3 else 'N/A',
                 'ptoTrabajoDesc': str(puestos_mapping.get(k.split('||')[3] if len(k.split('||')) > 3 else 'N/A', 'N/A')).capitalize() if puestos_mapping else 'N/A',
                 'cantidad': group_avisos[k]}
                for k in sorted(group_avisos.keys())
            ]
        },
        'resumenOrdenes': {
            'total': total_ordenes_count,
            'distribucion': [
                {'proceso': k.split('||')[0], 'grPlanif': k.split('||')[1],
                 'grPlanifPM': k.split('||')[2],
                 'ptoTrabajo': k.split('||')[3] if len(k.split('||')) > 3 else 'N/A',
                 'ptoTrabajoDesc': str(puestos_mapping.get(k.split('||')[3] if len(k.split('||')) > 3 else 'N/A', 'N/A')).capitalize() if puestos_mapping else 'N/A',
                 'cantidad': group_ordenes[k]}
                for k in sorted(group_ordenes.keys())
            ]
        },
        'trabajoPlanificado': {
            'grupos': cump_trabajo_planificado,
            'total': {
                'planificado': total_tp_planificado,
                'sinHr':       total_tp_sin_hr,
                'imprevistos': total_tp_imprevistos,
                'total':       total_tp_total,
                'cumplimiento': float(total_tp_cumplimiento_value
                                      if total_tp_cumplimiento_value is not None
                                      else total_tp_cumplimiento)
            }
        },
        'programaSemanal': {
            'grupos': cump_programa_semanal,
            'total': {
                'cumple':    total_prog_cumple,
                'noCumple':  total_prog_no_cumple,
                'total':     total_prog_total_ops,
                'cumplimiento': float(total_prog_cumplimiento_value
                                      if total_prog_cumplimiento_value is not None
                                      else total_prog_cumplimiento)
            }
        },
        'planMatriz': {
            'grupos': cump_plan_matriz,
            'total': {
                'cumple':    total_mtz_ejec,
                'noCumple':  total_mtz_totales - total_mtz_ejec,
                'total':     total_mtz_totales,
                'cumplimiento': float(total_mtz_cumplimiento)
            }
        }
    }
